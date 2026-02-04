import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_pdf import PdfPages
import os
import sys
import textwrap
import zipfile
import xml.etree.ElementTree as ET

class UTGMapVisualizer:
    def __init__(self, file_path, logo_path="./01_Logos/UTG_Logos/UTG_Logo_Gris_v3.0.png"):
        self.file_path = file_path
        self.logo_path = logo_path
        self.file_name = os.path.splitext(os.path.basename(file_path))[0]
        self.output_folder = "12_Licenciaturas_PDF+PNG"
        
        self.excel_data = None
        self.colors_map = {} 
        
        # Configuración visual
        self.box_width = 0.85
        self.box_height = 0.65
        self.v_spacing = 1.0
        self.stripe_colors = ['#FFFFFF', '#BBBBBB'] 

    def _get_contrast_color(self, hex_bg):
        if not hex_bg or not isinstance(hex_bg, str) or not hex_bg.startswith('#'):
            return 'black'
        try:
            hex_bg = hex_bg.lstrip('#')
            if len(hex_bg) != 6: return 'black'
            r, g, b = int(hex_bg[0:2], 16), int(hex_bg[2:4], 16), int(hex_bg[4:6], 16)
            if (r < 50 and g < 50 and b < 50):
                comp_r, comp_g, comp_b = 255 - r, 255 - g, 255 - b
                return f'#{comp_r:02x}{comp_g:02x}{comp_b:02x}'
            return 'black'
        except Exception:
            return 'black'

    def load_data(self):
        ext = os.path.splitext(self.file_path)[1].lower()
        try:
            if ext == '.ods':
                xls_file = pd.ExcelFile(self.file_path, engine='odf')
                self.sheets = xls_file.sheet_names
                self.data = {sheet: xls_file.parse(sheet, header=None) for sheet in self.sheets}
                print(f"   [ODS] Extrayendo colores de {self.file_name}...")
                self._extract_colors_ods()
            else:
                xls_file = pd.ExcelFile(self.file_path)
                self.sheets = xls_file.sheet_names
                self.data = {sheet: xls_file.parse(sheet, header=None) for sheet in self.sheets}
                self._extract_colors_xlsx()
        except Exception as e:
            print(f"Error crítico cargando {self.file_path}: {e}")
            self.sheets = []

    def _extract_colors_ods(self):
        """Versión DEFINITIVA (XML Parsing Nativo sin odfpy)."""
        ns = {
            'style': 'urn:oasis:names:tc:opendocument:xmlns:style:1.0',
            'fo': 'urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0',
            'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
            'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0'
        }
        style_map = {}
        try:
            with zipfile.ZipFile(self.file_path, 'r') as z:
                # 1. Leer estilos
                for xml_file in ['content.xml', 'styles.xml']:
                    if xml_file in z.namelist():
                        xml_content = z.read(xml_file)
                        tree = ET.fromstring(xml_content)
                        for style_node in tree.findall('.//style:style', ns):
                            style_name = style_node.get(f"{{{ns['style']}}}name")
                            cell_props = style_node.find('style:table-cell-properties', ns)
                            if cell_props is not None:
                                bg = cell_props.get(f"{{{ns['fo']}}}background-color")
                                if bg and bg != 'transparent' and bg != 'none':
                                    style_map[style_name] = bg

                # 2. Mapear celdas
                if 'content.xml' in z.namelist():
                    tree = ET.fromstring(z.read('content.xml'))
                    body = tree.find('office:body', ns)
                    spreadsheet = body.find('office:spreadsheet', ns) if body is not None else None
                    if spreadsheet is not None:
                        for table in spreadsheet.findall('table:table', ns):
                            sheet_name = table.get(f"{{{ns['table']}}}name")
                            row_idx = 0
                            for row in table.findall('table:table-row', ns):
                                rows_repeated = int(row.get(f"{{{ns['table']}}}number-rows-repeated") or 1)
                                if rows_repeated > 1000:
                                    row_idx += rows_repeated
                                    continue
                                col_idx = 0
                                for cell in row.findall('table:table-cell', ns):
                                    cols_repeated = int(cell.get(f"{{{ns['table']}}}number-columns-repeated") or 1)
                                    cell_style = cell.get(f"{{{ns['table']}}}style-name")
                                    if cell_style in style_map:
                                        color = style_map[cell_style]
                                        limit_r = min(rows_repeated, 500)
                                        for r in range(limit_r):
                                            for c in range(cols_repeated):
                                                self.colors_map[(sheet_name, row_idx + r, col_idx + c)] = color
                                    col_idx += cols_repeated
                                row_idx += rows_repeated
        except Exception as e:
            print(f"Error parseando XML nativo ODS: {e}")

    def _extract_colors_xlsx(self):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.fill and cell.fill.start_color:
                            color = cell.fill.start_color.rgb
                            if color and isinstance(color, str) and len(color) > 6: 
                                hex_c = '#' + color[2:] 
                                self.colors_map[(sheet_name, cell.row-1, cell.column-1)] = hex_c
        except Exception as e:
            print(f"Advertencia XLSX Colors: {e}")

    def _add_rounded_box(self, ax, x, y, text, bg_color=None, subtext=None):
        if bg_color is None: bg_color = '#E6E6E6'
        text_color = self._get_contrast_color(bg_color)
        wrapped_text = textwrap.fill(str(text), width=21, break_long_words=False)

        box = patches.FancyBboxPatch(
            (x - self.box_width/2, y - self.box_height/2),
            self.box_width, self.box_height,
            boxstyle="round,pad=0.05,rounding_size=0.1",
            linewidth=1, edgecolor='black', facecolor=bg_color, zorder=10
        )
        ax.add_patch(box)
        ax.text(x, y + 0.05, wrapped_text, ha='center', va='center', fontsize=6.5, 
                fontweight='bold', color=text_color, zorder=11)
        if subtext:
             ax.text(x, y - 0.20, subtext, ha='center', va='center', fontsize=5, 
                color=text_color, zorder=11, alpha=0.8)

    def _draw_logo(self, ax):
        if os.path.exists(self.logo_path):
            logo_img = plt.imread(self.logo_path)
            logo_ax = ax.inset_axes([0.01, 0.92, 0.08, 0.08])
            logo_ax.imshow(logo_img)
            logo_ax.axis('off')

    def render_main_map(self, sheet_name):
        df = self.data[sheet_name]
        rows, cols = df.shape
        width_fig = max(14, cols * 1.5)
        height_fig = max(8.5, rows * 1.2)
        
        fig, ax = plt.subplots(figsize=(width_fig, height_fig))
        ax.set_xlim(0, cols + 1.5)
        ax.set_ylim(0, rows + 2)
        ax.axis('off')
        
        # --- CAMBIO 1: Barras sin texto lateral ---
        for row_idx in range(rows):
            y_pos = (rows + 0.5) - (row_idx * self.v_spacing)
            rect = patches.Rectangle(
                (0, y_pos - 0.5), cols + 2, 1.0, 
                linewidth=0, facecolor=self.stripe_colors[row_idx % 2], zorder=0
            )
            ax.add_patch(rect)
            # SE ELIMINÓ: ax.text(...) que ponía "Sem 1", "Sem 2"...

        for col_idx in range(cols): 
            for row_idx in range(rows):
                cell_val = df.iloc[row_idx, col_idx]
                if pd.notna(cell_val) and str(cell_val).strip() != "":
                    x = col_idx + 1.5
                    y = (rows + 0.5) - (row_idx * self.v_spacing)
                    bg_color = self.colors_map.get((sheet_name, row_idx, col_idx), '#E6E6E6')
                    cr = "8.0 Cr" if "Optativa" not in str(cell_val) else "Variable"
                    self._add_rounded_box(ax, x, y, str(cell_val), bg_color=bg_color, subtext=cr)

        self._draw_logo(ax)
        plt.title(f"Mapa Curricular: {self.file_name}", fontsize=14, y=0.98)
        plt.tight_layout()
        return fig

    def render_specialization_map(self, sheet_name):
        df = self.data[sheet_name]
        rows, cols = df.shape
        
        needed_x_limit = max(10, (cols * 2.5) + 1)
        needed_y_limit = max(10, (rows * 1.5) + 2)
        
        fig_width = max(14, needed_x_limit * 1.2) 
        fig_height = max(8.5, needed_y_limit * 0.8)
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        ax.set_xlim(0, needed_x_limit)
        ax.set_ylim(0, needed_y_limit)
        ax.axis('off')
        
        # Título
        ax.text(needed_x_limit / 2, needed_y_limit - 0.5, sheet_name, 
                ha='center', va='center', fontsize=12, fontweight='bold')
        
        start_y = needed_y_limit - 2
        
        for row_idx in range(rows):
            sem_label = str(df.iloc[row_idx, 0]) if cols > 0 else ""
            
            # --- CAMBIO 2: Barras y Etiquetas en Especialidades ---
            y_center = start_y - (row_idx * 1.5)
            
            # Dibujar franja horizontal
            rect = patches.Rectangle(
                (0, y_center - 0.75), needed_x_limit, 1.5,
                linewidth=0, facecolor=self.stripe_colors[row_idx % 2], zorder=0
            )
            ax.add_patch(rect)
            
            # Etiqueta de semestre a la izquierda (Columna 0)
            if sem_label and sem_label.lower() != 'nan':
                 ax.text(0.8, y_center, f"Sem {sem_label}", 
                    ha='center', va='center', fontsize=9, fontweight='bold', color='#555')

            # Dibujar materias (Cols 1 en adelante)
            for col_idx in range(1, cols):
                cell_val = df.iloc[row_idx, col_idx]
                if pd.notna(cell_val) and str(cell_val).strip() != "":
                    x = 2 + (col_idx * 2.5) 
                    y = y_center
                    
                    bg_color = self.colors_map.get((sheet_name, row_idx, col_idx), '#DDA0DD')
                    
                    if col_idx > 1:
                        prev_x = 2 + ((col_idx - 1) * 2.5)
                        plt.plot([prev_x + (self.box_width/2), x - (self.box_width/2)], 
                                 [y, y], color='gray', lw=1, zorder=5)
                    
                    self._add_rounded_box(ax, x, y, str(cell_val), bg_color=bg_color, subtext=f"Sem {sem_label}")

        self._draw_logo(ax)
        plt.tight_layout()
        return fig

    def generate_outputs(self):
        self.load_data()
        if not hasattr(self, 'sheets') or not self.sheets: return

        os.makedirs(self.output_folder, exist_ok=True)
        pdf_path = os.path.join(self.output_folder, f"{self.file_name}.pdf")
        
        print(f"Generando salida en: {pdf_path}")
        
        with PdfPages(pdf_path) as pdf:
            if len(self.sheets) > 0:
                print(f"   Renderizando mapa principal...")
                fig_main = self.render_main_map(self.sheets[0])
                pdf.savefig(fig_main)
                png_path = os.path.join(self.output_folder, f"{self.file_name}_Main.png")
                fig_main.savefig(png_path, dpi=300, bbox_inches='tight')
                plt.close(fig_main)
            
            for sheet in self.sheets[1:]:
                print(f"   Renderizando especialidad: {sheet}")
                fig_spec = self.render_specialization_map(sheet)
                pdf.savefig(fig_spec)
                safe_sheet = "".join([c for c in sheet if c.isalnum() or c in (' ','-','_')]).strip()
                png_path = os.path.join(self.output_folder, f"{self.file_name}_{safe_sheet}.png")
                fig_spec.savefig(png_path, dpi=300, bbox_inches='tight')
                plt.close(fig_spec)

def main():
    carpeta_origen = "./12_Licenciaturas_BIS"
    lista_archivos = sys.argv[1:]
    
    if not lista_archivos:
        print("Advertencia: No se recibieron archivos desde Bash.")
        return

    for nombre_archivo in lista_archivos:
        ruta_completa = os.path.join(carpeta_origen, nombre_archivo)
        if os.path.exists(ruta_completa):
            print(f"\n--- Procesando: {nombre_archivo} ---")
            viz = UTGMapVisualizer(ruta_completa)
            viz.generate_outputs()
        else:
            print(f"Archivo no encontrado: {ruta_completa}")

if __name__ == "__main__":
    main()
